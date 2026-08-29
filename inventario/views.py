import io
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Producto, Categoria, MovimientoInventario
from .services import registrar_movimiento, calcular_metricas_inventario


def dashboard_view(request):
    """
    Vista principal estilo Cuenti / ERP.
    Muestra KPIs globales y la tabla de productos con análisis estocástico en tiempo real.
    """
    productos = Producto.objects.filter(activo=True).select_related('categoria')
    
    productos_con_metricas = []
    total_valor_inventario = Decimal('0.00')
    conteo_criticos = 0
    conteo_reorden = 0

    for prod in productos:
        metricas = calcular_metricas_inventario(prod)
        productos_con_metricas.append({
            'obj': prod,
            'metricas': metricas
        })
        total_valor_inventario += prod.stock_actual * prod.costo_unitario
        if metricas['estado'] == 'CRITICO':
            conteo_criticos += 1
        elif metricas['estado'] == 'REORDENAR':
            conteo_reorden += 1

    context = {
        'productos': productos_con_metricas,
        'total_productos': productos.count(),
        'total_valor_inventario': total_valor_inventario,
        'conteo_criticos': conteo_criticos,
        'conteo_reorden': conteo_reorden,
    }
    return render(request, 'inventario/dashboard.html', context)


def registrar_movimiento_view(request):
    """
    Controlador para registrar entradas, salidas y ajustes de stock en el Kárdex.
    """
    if request.method == 'POST':
        producto_id = request.POST.get('producto_id')
        tipo = request.POST.get('tipo')
        cantidad = Decimal(request.POST.get('cantidad', '0'))
        costo_str = request.POST.get('costo_unitario', '').strip()
        costo_unitario = Decimal(costo_str) if costo_str else None
        motivo = request.POST.get('motivo', '')

        producto = get_object_or_404(Producto, pk=producto_id)

        try:
            registrar_movimiento(
                producto=producto,
                tipo=tipo,
                cantidad=cantidad,
                costo_unitario=costo_unitario,
                motivo=motivo
            )
            messages.success(request, f"Movimiento de {tipo} registrado exitosamente para {producto.nombre}.")
        except Exception as e:
            messages.error(request, f"Error al registrar movimiento: {str(e)}")

        return redirect('dashboard')

    # Si es GET, mostramos el listado de productos para el selector
    productos = Producto.objects.filter(activo=True)
    return render(request, 'inventario/registrar_movimiento.html', {'productos': productos})


def exportar_kardex_excel_view(request, producto_id=None):
    """
    Genera un archivo Excel (.xlsx) con el historial de Kárdex profesional usando OpenPyXL.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex de Inventario"

    # Estilos profesionales
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Azul oscuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "Fecha / Hora", "SKU", "Producto", "Tipo Movimiento", 
        "Cantidad", "Costo Unitario", "Stock Anterior", "Saldo Resultante", "Motivo"
    ]

    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filtrado de movimientos
    if producto_id:
        movimientos = MovimientoInventario.objects.filter(producto_id=producto_id).select_related('producto')
    else:
        movimientos = MovimientoInventario.objects.all().select_related('producto')[:500]

    for mov in movimientos:
        row = [
            mov.fecha.strftime("%Y-%m-%d %H:%M"),
            mov.producto.sku,
            mov.producto.nombre,
            mov.tipo,
            float(mov.cantidad),
            float(mov.costo_unitario),
            float(mov.stock_anterior),
            float(mov.stock_resultante),
            mov.motivo or "-"
        ]
        ws.append(row)
        current_row = ws.max_row
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border_thin
            if col_num in [5, 6, 7, 8]: # Columnas numéricas
                cell.alignment = Alignment(horizontal="right")

    # Autoajuste del ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Kardex_Stokify_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

